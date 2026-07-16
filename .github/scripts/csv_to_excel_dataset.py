"""
csv_to_excel_dataset.py

Converts one CDISC Open Rules test-case `data/` folder (.env, _datasets.csv,
_variables.csv, and one CSV per dataset) into the single-workbook Excel
format required by Verisian's ExcelDataService
(cdisc_rules_engine/services/data_services/excel_data_service.py):

    - Exactly one .xlsx file.
    - A "Datasets" sheet with columns Filename, Label.
    - One sheet per dataset, sheet name == the Filename value (with .xpt
      appended if not already present, since Verisian's engine expects
      dataset filenames to carry the .xpt extension).
    - Each dataset sheet's first 4 rows (no header row skipped) are:
        row 1: variable names
        row 2: variable labels
        row 3: variable types   -- MUST stay exactly "Char"/"Num" as written
                                    in _variables.csv. ExcelDataService reads
                                    these case-sensitively
                                    ({"Char": str, "Num": float, ...}) and
                                    silently falls back to str for anything
                                    that doesn't match, so do NOT lowercase.
        row 4: variable lengths
      followed by the actual data from row 5 onward.

Also returns the parsed .env values (PRODUCT, VERSION, SUBSTANDARD, ...) so
the caller can build the `core.py validate` CLI arguments.
"""

import csv
import os
import re
from collections import defaultdict
from typing import Dict, List, Tuple

from openpyxl import Workbook

REQUIRED_FILES = ("_datasets.csv", "_variables.csv")


class ConversionError(Exception):
    pass


def find_env_file(data_dir: str) -> str:
    """
    Locate the .env file in a test case's data/ folder. Matches an exact
    '.env' filename, but also tolerates a file merely ending in '.env' in
    case a differently-named variant ever shows up.
    """
    for name in os.listdir(data_dir):
        if name == ".env" or name.endswith(".env"):
            return os.path.join(data_dir, name)
    raise ConversionError(f"No .env file found in {data_dir}")


def read_env(path: str) -> Dict[str, str]:
    env = {}
    with open(path, "r", encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key.strip().upper()] = value.strip()
    return env


def read_csv_rows(path: str) -> List[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_csv_raw(path: str) -> Tuple[List[str], List[List[str]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def base_name(filename: str) -> str:
    return re.sub(r"\.[A-Za-z0-9]+$", "", filename.strip()).lower()


def ensure_xpt_extension(filename: str) -> str:
    name = filename.strip()
    if not name.lower().endswith(".xpt"):
        name = f"{name}.xpt"
    return name


def to_number_if_possible(value):
    if value is None:
        return None
    v = str(value).strip()
    if v == "":
        return None
    try:
        if re.fullmatch(r"[+-]?\d+", v):
            return int(v)
        return float(v)
    except ValueError:
        return value


def check_required_files(data_dir: str) -> List[str]:
    missing = [name for name in REQUIRED_FILES if not os.path.isfile(os.path.join(data_dir, name))]
    try:
        find_env_file(data_dir)
    except ConversionError:
        missing.append(".env")
    return missing


def convert_test_case_to_excel(data_dir: str, output_xlsx_path: str) -> Dict[str, str]:
    """
    Converts a single test case's data/ folder into one .xlsx workbook at
    output_xlsx_path, matching Verisian's ExcelDataService expectations.

    Returns the parsed .env dict (e.g. {"PRODUCT": "SDTMIG", "VERSION": "3-3"}).
    Raises ConversionError on any missing/invalid required input.
    """
    missing = check_required_files(data_dir)
    if missing:
        raise ConversionError(f"missing required file(s) in {data_dir}: {', '.join(missing)}")

    env = read_env(find_env_file(data_dir))
    if "PRODUCT" not in env or "VERSION" not in env:
        raise ConversionError(f".env in {data_dir} must define PRODUCT and VERSION")

    dataset_rows = read_csv_rows(os.path.join(data_dir, "_datasets.csv"))
    if not dataset_rows:
        raise ConversionError(f"_datasets.csv in {data_dir} has no rows")
    for row in dataset_rows:
        row["Filename"] = ensure_xpt_extension(row["Filename"])

    variable_rows = read_csv_rows(os.path.join(data_dir, "_variables.csv"))
    variables_by_dataset = defaultdict(list)
    for row in variable_rows:
        variables_by_dataset[row["dataset"].strip().lower()].append(row)

    wb = Workbook()
    # Remove the default sheet; we'll add "Datasets" explicitly so it's first.
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_ds = wb.create_sheet("Datasets")
    ws_ds.append(["Filename", "Label"])
    for row in dataset_rows:
        ws_ds.append([row["Filename"], row["Label"]])

    for row in dataset_rows:
        filename = row["Filename"]
        base = base_name(filename)

        var_rows = variables_by_dataset.get(base, [])
        if not var_rows:
            # Fallback: longest dataset-name prefix match (handles split
            # datasets, e.g. variables listed under "ec" but files "ecaa"/"ecbb")
            candidates = [k for k in variables_by_dataset if base.startswith(k)]
            if candidates:
                var_rows = variables_by_dataset[max(candidates, key=len)]
        if not var_rows:
            raise ConversionError(f"No variable metadata in _variables.csv for dataset '{filename}' in {data_dir}")

        sheet_name = filename[:31]
        ws = wb.create_sheet(sheet_name)

        var_names = [v["variable"] for v in var_rows]
        var_labels = [v["label"] for v in var_rows]
        # IMPORTANT: keep type exactly as written (e.g. "Char"/"Num") —
        # ExcelDataService matches these case-sensitively.
        var_types = [v["type"].strip() for v in var_rows]
        var_lengths = [to_number_if_possible(v["length"]) for v in var_rows]

        ws.append(var_names)
        ws.append(var_labels)
        ws.append(var_types)
        ws.append(var_lengths)

        src_path = os.path.join(data_dir, f"{base}.csv")
        if not os.path.isfile(src_path):
            raise ConversionError(f"No source data CSV found for '{filename}' (expected '{base}.csv') in {data_dir}")

        header, data_rows = read_csv_raw(src_path)
        header_index = {name: i for i, name in enumerate(header)}
        type_by_var = {v["variable"]: v["type"].strip() for v in var_rows}

        for data_row in data_rows:
            out_row = []
            for col in var_names:
                idx = header_index.get(col)
                raw_val = data_row[idx] if idx is not None and idx < len(data_row) else ""
                if type_by_var.get(col) == "Num":
                    out_row.append(to_number_if_possible(raw_val))
                else:
                    out_row.append(raw_val if raw_val != "" else None)
            ws.append(out_row)

    wb.save(output_xlsx_path)
    return env
